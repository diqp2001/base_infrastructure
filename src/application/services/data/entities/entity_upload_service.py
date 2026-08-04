"""
EntityUploadService — Excel template generation and upload processing for domain entities.

Supports: Continent, Country, Exchange, Sector, Industry, Currency, Index,
          CompanyShare, Model, Universe, BacktestFactor.
"""

import io
from datetime import date, datetime
from typing import Any, Dict, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Schema definitions
# ---------------------------------------------------------------------------

ENTITY_SCHEMAS: Dict[str, Dict[str, Any]] = {
    "Continent": {
        "columns": ["name"],
        "required": ["name"],
        "example_row": {"name": "Europe"},
        "description": "name: unique continent name (e.g. Europe, Asia, North America)",
    },
    "Country": {
        "columns": ["name", "iso_code", "continent_name"],
        "required": ["name", "iso_code", "continent_name"],
        "example_row": {"name": "France", "iso_code": "FR", "continent_name": "Europe"},
        "description": (
            "name: country name; iso_code: 2-letter ISO code; "
            "continent_name: must match an existing Continent name"
        ),
    },
    "Exchange": {
        "columns": ["name", "legal_name", "country_name", "start_date", "end_date"],
        "required": ["name", "legal_name", "country_name", "start_date"],
        "example_row": {
            "name": "NYSE",
            "legal_name": "New York Stock Exchange",
            "country_name": "United States",
            "start_date": "1817-01-01",
            "end_date": "",
        },
        "description": (
            "name: short exchange name; legal_name: full legal name; "
            "country_name: must match existing Country; "
            "start_date/end_date: YYYY-MM-DD (end_date optional)"
        ),
    },
    "Sector": {
        "columns": ["name", "description"],
        "required": ["name"],
        "example_row": {
            "name": "Technology",
            "description": "Technology companies and services",
        },
        "description": "name: unique sector name; description: optional text",
    },
    "Industry": {
        "columns": ["name", "sector_name", "description"],
        "required": ["name", "sector_name"],
        "example_row": {
            "name": "Software",
            "sector_name": "Technology",
            "description": "Software development and services",
        },
        "description": (
            "name: industry name; sector_name: must match an existing Sector; "
            "description: optional text"
        ),
    },
    "Currency": {
        "columns": ["name", "symbol", "country_name", "start_date", "end_date"],
        "required": ["name", "symbol"],
        "example_row": {
            "name": "Euro",
            "symbol": "EUR",
            "country_name": "",
            "start_date": "1999-01-01",
            "end_date": "",
        },
        "description": (
            "name: currency name; symbol: ISO/ticker code (e.g. USD, EUR); "
            "country_name: optional, must match Country if provided; "
            "start_date/end_date: optional YYYY-MM-DD"
        ),
    },
    "Index": {
        "columns": ["name", "symbol", "currency_symbol", "start_date", "end_date"],
        "required": ["name", "symbol"],
        "example_row": {
            "name": "S&P 500",
            "symbol": "SPX",
            "currency_symbol": "USD",
            "start_date": "1957-03-04",
            "end_date": "",
        },
        "description": (
            "name: index name; symbol: ticker; "
            "currency_symbol: optional, must match Currency.symbol; "
            "start_date/end_date: optional YYYY-MM-DD"
        ),
    },
    "CompanyShare": {
        "columns": [
            "name",
            "symbol",
            "currency_symbol",
            "exchange_name",
            "company_id",
            "start_date",
            "end_date",
        ],
        "required": ["symbol", "currency_symbol", "exchange_name"],
        "example_row": {
            "name": "Apple Inc.",
            "symbol": "AAPL",
            "currency_symbol": "USD",
            "exchange_name": "NASDAQ",
            "company_id": "",
            "start_date": "1980-12-12",
            "end_date": "",
        },
        "description": (
            "symbol: stock ticker (required); name: company display name (optional); "
            "currency_symbol: must match Currency.symbol; "
            "exchange_name: must match Exchange.name; "
            "company_id: optional numeric ID of related Company (defaults to 0); "
            "start_date/end_date: optional YYYY-MM-DD"
        ),
    },
    "Model": {
        "columns": ["name"],
        "required": ["name"],
        "example_row": {"name": "momentum_v1"},
        "description": "name: unique backtest model / algorithm name",
    },
    "Universe": {
        "columns": ["name", "creation_date", "description"],
        "required": ["name", "creation_date"],
        "example_row": {
            "name": "SP500_2024",
            "creation_date": "2024-01-01",
            "description": "S&P 500 universe for 2024",
        },
        "description": (
            "name: unique universe name; creation_date: YYYY-MM-DD; description: optional"
        ),
    },
    "BacktestFactor": {
        "columns": [
            "name",
            "group",
            "subgroup",
            "frequency",
            "data_type",
            "source",
            "definition",
        ],
        "required": ["name", "group"],
        "example_row": {
            "name": "momentum_20d",
            "group": "momentum",
            "subgroup": "daily",
            "frequency": "1d",
            "data_type": "numeric",
            "source": "calculated",
            "definition": "20-day momentum factor",
        },
        "description": (
            "name: factor name; group: factor group (e.g. momentum, return, value, price); "
            "subgroup/frequency/data_type/source/definition: optional"
        ),
    },
    "FactorValue": {
        "columns": [
            "entity_type",
            "entity_name",
            "factor_name",
            "date",
            "value",
            "currency_name",
        ],
        "required": ["entity_type", "entity_name", "factor_name", "date", "value", "currency_name"],
        "example_row": {
            "entity_type": "CompanyShare",
            "entity_name": "Apple Inc.",
            "factor_name": "company_share_mid_price",
            "date": "2024-01-15",
            "value": "182.50",
            "currency_name": "US Dollar",
        },
        "description": (
            "entity_type: domain entity class name (e.g. CompanyShare, Currency, Index); "
            "entity_name: name as stored in DB; "
            "factor_name: exact factor name in DB; "
            "date: YYYY-MM-DD; "
            "value: numeric string; "
            "currency_name: name of the currency the value is expressed in"
        ),
    },
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_date(val) -> Optional[date]:
    """Parse a date from a cell value (str, datetime, date, or NaN/None)."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _str_or_none(val) -> Optional[str]:
    """Return stripped string or None if blank / NaN."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    return s if s else None


def _next_id(repo) -> int:
    """
    Get the next available sequential ID from a repository.
    Tries the base _get_next_available_id() first; falls back to a direct query.
    """
    try:
        model_class = repo.model_class
        result = (
            repo.session.query(model_class.id)
            .order_by(model_class.id.desc())
            .first()
        )
        if result is not None:
            # SQLAlchemy scalar queries return Row; support both tuple and attr access
            val = result[0] if hasattr(result, "__getitem__") else result.id
            return int(val) + 1
        return 1
    except Exception as exc:
        print(f"Warning: _next_id fallback (1) due to: {exc}")
        return 1


def _lookup_by_symbol(repo, symbol: str):
    """
    Look up a financial entity by its symbol column.
    Falls back to get_by_name if symbol lookup fails.
    """
    try:
        model_class = repo.model_class
        model = (
            repo.session.query(model_class)
            .filter(model_class.symbol == symbol)
            .first()
        )
        if model is not None:
            return repo._to_entity(model)
    except Exception:
        pass
    # Fallback: try name == symbol
    try:
        return repo.get_by_name(symbol)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def generate_template(entity_type: str) -> io.BytesIO:
    """
    Generate an Excel workbook for the given entity type.

    Sheet 'Data'        — bold blue header row + one example data row.
    Sheet 'Instructions'— entity description + per-column notes.

    Returns a seeked BytesIO ready for send_file.
    """
    if entity_type not in ENTITY_SCHEMAS:
        raise ValueError(f"Unknown entity type: {entity_type!r}")

    schema = ENTITY_SCHEMAS[entity_type]
    columns: list = schema["columns"]
    required: list = schema["required"]
    example_row: dict = schema["example_row"]
    description: str = schema["description"]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ──────────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        col_letter = cell.column_letter
        ws_data.column_dimensions[col_letter].width = max(len(col_name) + 6, 18)

    for col_idx, col_name in enumerate(columns, start=1):
        ws_data.cell(row=2, column=col_idx, value=example_row.get(col_name, ""))

    # ── Sheet 2: Instructions ──────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.column_dimensions["A"].width = 22
    ws_inst.column_dimensions["B"].width = 12
    ws_inst.column_dimensions["C"].width = 65

    title_cell = ws_inst.cell(row=1, column=1, value=f"{entity_type} — Upload Instructions")
    title_cell.font = Font(bold=True, size=13)

    ws_inst.cell(row=3, column=1, value="Entity description")
    ws_inst.cell(row=3, column=1).font = Font(bold=True)
    ws_inst.cell(row=3, column=2, value=description)

    ws_inst.cell(row=5, column=1, value="Column").font = Font(bold=True)
    ws_inst.cell(row=5, column=2, value="Required?").font = Font(bold=True)
    ws_inst.cell(row=5, column=3, value="Example value").font = Font(bold=True)

    for row_idx, col_name in enumerate(columns, start=6):
        ws_inst.cell(row=row_idx, column=1, value=col_name)
        ws_inst.cell(
            row=row_idx,
            column=2,
            value="YES" if col_name in required else "no",
        )
        ws_inst.cell(row=row_idx, column=3, value=str(example_row.get(col_name, "")))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Upload processing
# ---------------------------------------------------------------------------


def process_upload(entity_type: str, df: pd.DataFrame, session) -> dict:
    """
    Process rows from an uploaded DataFrame and persist domain entities.

    Args:
        entity_type: Key into ENTITY_SCHEMAS (e.g. 'Continent').
        df:          DataFrame read from the 'Data' sheet of the uploaded file.
        session:     SQLAlchemy session.

    Returns:
        {'processed': N, 'errors': [{'row': int, 'error': str}, ...]}
    """
    if entity_type not in ENTITY_SCHEMAS:
        return {
            "processed": 0,
            "errors": [{"row": 0, "error": f"Unknown entity type: {entity_type!r}"}],
        }

    # Lazy import to avoid triggering SQLAlchemy model registration at module load time
    from src.infrastructure.repositories.repository_factory import RepositoryFactory

    factory = RepositoryFactory(session)
    handlers = _build_handlers(factory, session)

    if entity_type not in handlers:
        return {
            "processed": 0,
            "errors": [
                {
                    "row": 0,
                    "error": f"No upload handler defined for entity type: {entity_type!r}",
                }
            ],
        }

    handler = handlers[entity_type]
    processed = 0
    errors = []

    for df_idx, row in df.iterrows():
        row_num = int(df_idx) + 2  # Excel row number (1=header, 2=first data row)
        try:
            result = handler(row)
            if result is not None:
                processed += 1
            else:
                errors.append(
                    {
                        "row": row_num,
                        "error": (
                            "Entity creation returned None — check that all FK references "
                            "exist and required fields are non-empty."
                        ),
                    }
                )
        except Exception as exc:
            errors.append({"row": row_num, "error": str(exc)})
            try:
                session.rollback()
            except Exception:
                pass

    return {"processed": processed, "errors": errors}


# ---------------------------------------------------------------------------
# Per-entity handlers
# ---------------------------------------------------------------------------


def _build_handlers(factory, session) -> dict:  # noqa: C901
    """Return a dict of entity_type → callable(row) → entity | None."""

    # ── Continent ────────────────────────────────────────────────────────
    def _continent(row):
        name = _str_or_none(row.get("name"))
        if not name:
            raise ValueError("'name' is required")
        repo = factory.get_local_repository("Continent")
        return repo._create_or_get(name=name)

    # ── Country ──────────────────────────────────────────────────────────
    def _country(row):
        name = _str_or_none(row.get("name"))
        iso_code = _str_or_none(row.get("iso_code"))
        continent_name = _str_or_none(row.get("continent_name"))
        if not name or not iso_code or not continent_name:
            raise ValueError("'name', 'iso_code', and 'continent_name' are required")

        continent_repo = factory.get_local_repository("Continent")
        continent = continent_repo.get_by_name(continent_name)
        if continent is None:
            raise ValueError(f"Continent '{continent_name}' not found")

        country_repo = factory.get_local_repository("Country")
        existing = country_repo.get_by_name(name)
        if existing:
            return existing

        from src.domain.entities.country import Country

        entity = Country(id=_next_id(country_repo), name=name, iso_code=iso_code, continent_id=continent.id)
        orm_model = country_repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return country_repo._to_entity(orm_model)

    # ── Exchange ─────────────────────────────────────────────────────────
    def _exchange(row):
        name = _str_or_none(row.get("name"))
        legal_name = _str_or_none(row.get("legal_name"))
        country_name = _str_or_none(row.get("country_name"))
        start_date = _parse_date(row.get("start_date"))
        end_date = _parse_date(row.get("end_date"))

        if not name or not legal_name or not country_name or start_date is None:
            raise ValueError("'name', 'legal_name', 'country_name', and 'start_date' are required")

        country_repo = factory.get_local_repository("Country")
        country = country_repo.get_by_name(country_name)
        if country is None:
            raise ValueError(f"Country '{country_name}' not found")

        exchange_repo = factory.get_local_repository("Exchange")
        existing = exchange_repo.get_by_name(name)
        if existing:
            return existing

        from src.domain.entities.finance.exchange import Exchange

        entity = Exchange(
            id=_next_id(exchange_repo),
            name=name,
            legal_name=legal_name,
            country_id=country.id,
            start_date=start_date,
            end_date=end_date,
        )
        orm_model = exchange_repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return exchange_repo._to_entity(orm_model)

    # ── Sector ───────────────────────────────────────────────────────────
    def _sector(row):
        name = _str_or_none(row.get("name"))
        description = _str_or_none(row.get("description"))
        if not name:
            raise ValueError("'name' is required")
        repo = factory.get_local_repository("Sector")
        return repo._create_or_get(name=name, description=description)

    # ── Industry ─────────────────────────────────────────────────────────
    def _industry(row):
        name = _str_or_none(row.get("name"))
        sector_name = _str_or_none(row.get("sector_name"))
        description = _str_or_none(row.get("description"))

        if not name or not sector_name:
            raise ValueError("'name' and 'sector_name' are required")

        sector_repo = factory.get_local_repository("Sector")
        sector = sector_repo.get_by_name(sector_name)
        if sector is None:
            raise ValueError(f"Sector '{sector_name}' not found")

        industry_repo = factory.get_local_repository("Industry")
        existing = industry_repo.get_by_name(name)
        if existing:
            return existing

        from src.domain.entities.industry import Industry

        entity = Industry(id=_next_id(industry_repo), name=name, sector_id=sector.id, description=description or "")
        orm_model = industry_repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return industry_repo._to_entity(orm_model)

    # ── Currency ─────────────────────────────────────────────────────────
    def _currency(row):
        name = _str_or_none(row.get("name"))
        symbol = _str_or_none(row.get("symbol"))
        country_name = _str_or_none(row.get("country_name"))
        start_date = _parse_date(row.get("start_date"))
        end_date = _parse_date(row.get("end_date"))

        if not name or not symbol:
            raise ValueError("'name' and 'symbol' are required")

        currency_repo = factory.get_local_repository("Currency")
        existing = currency_repo.get_by_name(name)
        if existing:
            return existing

        country_id = None
        if country_name:
            country_repo = factory.get_local_repository("Country")
            country = country_repo.get_by_name(country_name)
            if country is None:
                raise ValueError(f"Country '{country_name}' not found")
            country_id = country.id

        from src.domain.entities.finance.financial_assets.currency import Currency

        entity = Currency(
            id=_next_id(currency_repo),
            name=name,
            symbol=symbol,
            country_id=country_id,
            start_date=start_date,
            end_date=end_date,
        )
        orm_model = currency_repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return currency_repo._to_entity(orm_model)

    # ── Index ────────────────────────────────────────────────────────────
    def _index(row):
        name = _str_or_none(row.get("name"))
        symbol = _str_or_none(row.get("symbol"))
        currency_symbol = _str_or_none(row.get("currency_symbol"))
        start_date = _parse_date(row.get("start_date"))
        end_date = _parse_date(row.get("end_date"))

        if not name or not symbol:
            raise ValueError("'name' and 'symbol' are required")

        index_repo = factory.get_local_repository("Index")
        existing = index_repo.get_by_name(name)
        if existing:
            return existing

        currency_id = None
        if currency_symbol:
            currency_repo = factory.get_local_repository("Currency")
            currency = _lookup_by_symbol(currency_repo, currency_symbol)
            if currency is None:
                raise ValueError(f"Currency with symbol '{currency_symbol}' not found")
            currency_id = currency.id

        from src.domain.entities.finance.financial_assets.index.index import Index

        entity = Index(
            id=_next_id(index_repo),
            name=name,
            symbol=symbol,
            currency_id=currency_id,
            start_date=start_date,
            end_date=end_date,
        )
        orm_model = index_repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return index_repo._to_entity(orm_model)

    # ── CompanyShare ─────────────────────────────────────────────────────
    def _company_share(row):
        name = _str_or_none(row.get("name"))
        symbol = _str_or_none(row.get("symbol"))
        currency_symbol = _str_or_none(row.get("currency_symbol"))
        exchange_name = _str_or_none(row.get("exchange_name"))
        company_id_raw = row.get("company_id")
        start_date = _parse_date(row.get("start_date"))
        end_date = _parse_date(row.get("end_date"))

        if not symbol or not currency_symbol or not exchange_name:
            raise ValueError("'symbol', 'currency_symbol', and 'exchange_name' are required")

        # Resolve company_id (defaults to 0 when omitted)
        try:
            if pd.isna(company_id_raw) or company_id_raw is None or str(company_id_raw).strip() == "":
                company_id = 0
            else:
                company_id = int(float(str(company_id_raw).strip()))
        except Exception:
            company_id = 0

        currency_repo = factory.get_local_repository("Currency")
        currency = _lookup_by_symbol(currency_repo, currency_symbol)
        if currency is None:
            raise ValueError(f"Currency with symbol '{currency_symbol}' not found")

        exchange_repo = factory.get_local_repository("Exchange")
        exchange = exchange_repo.get_by_name(exchange_name)
        if exchange is None:
            raise ValueError(f"Exchange '{exchange_name}' not found")

        share_repo = factory.get_local_repository("CompanyShare")
        # Check existence by symbol
        existing = _lookup_by_symbol(share_repo, symbol)
        if existing:
            return existing

        from src.domain.entities.finance.financial_assets.share.company_share.company_share import CompanyShare

        entity = CompanyShare(
            id=_next_id(share_repo),
            symbol=symbol,
            name=name,
            currency_id=currency.id,
            exchange_id=exchange.id,
            company_id=company_id,
            start_date=start_date,
            end_date=end_date,
        )
        orm_model = share_repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return share_repo._to_entity(orm_model)

    # ── Model ────────────────────────────────────────────────────────────
    def _model(row):
        name = _str_or_none(row.get("name"))
        if not name:
            raise ValueError("'name' is required")
        repo = factory.get_local_repository("Model")
        existing = repo.get_by_name(name)
        if existing:
            return existing
        from src.domain.entities.backtest.model import Model

        entity = Model(id=_next_id(repo), name=name)
        orm_model = repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return repo._to_entity(orm_model)

    # ── Universe ─────────────────────────────────────────────────────────
    def _universe(row):
        name = _str_or_none(row.get("name"))
        creation_date = _parse_date(row.get("creation_date"))
        description = _str_or_none(row.get("description"))

        if not name or creation_date is None:
            raise ValueError("'name' and 'creation_date' are required")

        repo = factory.get_local_repository("Universe")
        existing = repo.get_by_name(name)
        if existing:
            return existing

        from src.domain.entities.backtest.universe import Universe

        entity = Universe(id=_next_id(repo), name=name, creation_date=creation_date, description=description)
        orm_model = repo._to_model(entity)
        session.add(orm_model)
        session.commit()
        return repo._to_entity(orm_model)

    # ── BacktestFactor ───────────────────────────────────────────────────
    def _backtest_factor(row):
        name = _str_or_none(row.get("name"))
        group = _str_or_none(row.get("group"))
        if not name or not group:
            raise ValueError("'name' and 'group' are required")
        repo = factory.get_local_repository("BacktestFactor")
        return repo._create_or_get(
            entity_cls=None,
            primary_key=name,
            group=group,
            subgroup=_str_or_none(row.get("subgroup")),
            frequency=_str_or_none(row.get("frequency")),
            data_type=_str_or_none(row.get("data_type")),
            source=_str_or_none(row.get("source")),
            definition=_str_or_none(row.get("definition")),
        )

    # ── FactorValue ──────────────────────────────────────────────────────
    def _factor_value(row):
        entity_type = _str_or_none(row.get("entity_type"))
        entity_name = _str_or_none(row.get("entity_name"))
        factor_name = _str_or_none(row.get("factor_name"))
        fv_date = _parse_date(row.get("date"))
        value = _str_or_none(row.get("value"))
        currency_name = _str_or_none(row.get("currency_name"))

        if not all([entity_type, entity_name, factor_name, fv_date, value, currency_name]):
            raise ValueError(
                "All columns are required: entity_type, entity_name, factor_name, date, value, currency_name"
            )

        # Resolve entity_id from the repo keyed by entity_type
        entity_repo = factory.get_local_repository(entity_type)
        if entity_repo is None:
            raise ValueError(f"No repository found for entity_type '{entity_type}'")
        entity = entity_repo.get_by_name(entity_name)
        if entity is None:
            raise ValueError(f"Entity '{entity_name}' of type '{entity_type}' not found")

        # Resolve factor_id from the FactorValue repo's factor lookup
        factor_repo = factory.get_local_repository("Factor")
        from src.infrastructure.models.factor.factor import FactorModel
        factor_orm = (
            session.query(FactorModel).filter(FactorModel.name == factor_name).first()
        )
        if factor_orm is None:
            raise ValueError(f"Factor '{factor_name}' not found — create it first")

        # Resolve currency_id
        currency_repo = factory.get_local_repository("Currency")
        currency = currency_repo.get_by_name(currency_name)
        if currency is None:
            raise ValueError(f"Currency '{currency_name}' not found")

        # Persist FactorValue via ORM model directly
        from src.infrastructure.models.factor.factor_value import FactorValueModel
        from datetime import datetime as _dt, timezone as _tz
        fv_datetime = _dt.combine(fv_date, _dt.min.time()).replace(tzinfo=_tz.utc)

        existing = (
            session.query(FactorValueModel)
            .filter(
                FactorValueModel.factor_id == factor_orm.id,
                FactorValueModel.entity_id == entity.id,
                FactorValueModel.entity_type == entity_type,
                FactorValueModel.date == fv_datetime,
            )
            .first()
        )
        if existing:
            return existing

        orm_fv = FactorValueModel(
            factor_id=factor_orm.id,
            entity_id=entity.id,
            entity_type=entity_type,
            date=fv_datetime,
            value=str(value),
            currency_id=currency.id,
        )
        session.add(orm_fv)
        session.commit()
        return orm_fv

    return {
        "Continent": _continent,
        "Country": _country,
        "Exchange": _exchange,
        "Sector": _sector,
        "Industry": _industry,
        "Currency": _currency,
        "Index": _index,
        "CompanyShare": _company_share,
        "Model": _model,
        "Universe": _universe,
        "BacktestFactor": _backtest_factor,
        "FactorValue": _factor_value,
    }
